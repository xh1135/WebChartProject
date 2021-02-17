# from the flask package, import the Flask class
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, abort
import requests
import json
import sqlite3
from werkzeug.utils import secure_filename
import os
import sys
import unicodedata
from sqlalchemy import func

import uuid
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

# internal modules
from model import db, OpplanData, MyFormData, Profile, Restaurant, Tables, Employee
from forms import RegisterForm, LoginForm, UploadForm, OpplanForm, FormUtil, ProfileForm, TableForm, EmployeeForm


# instantiate Flask
app = Flask(__name__)
app.config['SECRET_KEY'] = 'mysecretkey'
# app.config['UPLOAD_FOLDER'] = os.path.dirname(sys.argv[0]) + "uploads"
app.config['UPLOAD_FOLDER'] = 'static' + os.path.sep + "uploads"

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///mydb.db'
# app.config['SQLALCHEMY_DATABASE_URI'] = 'oracle://MASY_IB48:MASY_IB48@localhost:1522/app12c'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

app.config['TOASTR_TIMEOUT'] = 1000

# db = SQLAlchemy(app)
# initalize app with database
db.init_app(app)


@app.before_first_request
def before_first_request_func():
    db.create_all()


@app.route("/")
def index():

    url = 'https://jsonplaceholder.typicode.com/todos'
    with requests.get(url) as response:
        data = json.loads(response.content)

    return render_template("home.html", title="Home", todos=data)


@app.route("/about")
def about():
    todos = []

    url = 'https://jsonplaceholder.typicode.com/todos'
    with requests.get(url) as response:
        data = json.loads(response.content)

    return render_template("about.html", title="About", todos=data)


@app.route("/shop")
def shop():
    return render_template("shop.html", title="Shop")


@app.route("/register", methods=['Get', 'Post'])
def register():
    form = RegisterForm()
    if form.validate_on_submit():
        try:
            user = User(form.name.data, form.email.data, form.password.data)

            with sqlite3.connect(DATABASE) as con:
                cur = con.cursor()
                cur.execute("INSERT INTO user (name,email,password) VALUES (?,?,?)",
                            (user.name, user.email, user.password))
                con.commit()
                flash('Registered Successfully!', 'success')
                return redirect(url_for('login'))
        except Exception as e:
            con.rollback()
            flash(f'Unknow error!\n{str(e)}', 'danger')

    return render_template('register.html', title='Register', form=form)


@app.route("/login", methods=['Get', 'Post'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        try:
            with sqlite3.connect(DATABASE) as con:
                cur = con.cursor()
                cur.execute("SELECT * FROM user WHERE email = ? and password = ?",
                            [form.email.data, form.password.data])
                rows = cur.fetchall()
                print(rows)
                row = rows[0] if rows else None
                if row:
                    flash('Login Successfully!', 'success')
                    return redirect(url_for('shop'))
                else:
                    flash(
                        'Login Unsuccessful. Please check email and password', 'danger')
        except Exception as e:
            con.rollback()
            flash(f'Unknow error!\n{str(e)}', 'danger')
    return render_template('login.html', title='Login', form=form)


@app.route("/upload", methods=['Get', 'Post'])
def upload():
    form = UploadForm()
    # print("method", request.method)

    my_form = MyFormData()

    if form.validate_on_submit():
        my_form.first_name = request.form['first_name']
        # form.first_name.errors.append("Invalid Name")

        my_form.last_name = request.form['last_name']
        my_form.email = request.form['email']
        my_form.password = request.form['password']

        my_form.comments = request.form['comments']
        my_form.my_date = request.form['my_date']
        my_form.my_date_range = request.form.getlist('my_date_range')
        my_form.multiple_button = request.form.getlist('multiple_button')
        my_form.single_button = request.form.get('single_button')

        my_form.single_state = request.form.get('single_state')
        my_form.multiple_states = request.form.getlist('multiple_states')

        my_form.toggle_switch = request.form.get('toggle_switch')
        # my_form.custom_range = request.form.get('custom_range')

        my_form.single_option = request.form.get('single_option')
        my_form.multiple_options = request.form.getlist('multiple_options')

        my_form.radio1 = request.form.get('radio1')
        my_form.check_list = request.form.getlist('check_list')

        # request.form['my_date_range']
        print("*****")
        print(my_form.check_list)
        print("*****")
        #
        # file = request.files['file_photo']
        # filename = secure_filename(file.filename)
        # print("path", os.path.join(app.config['UPLOAD_FOLDER']))
        # file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
    return render_template('upload.html', form=form, my_form=my_form)


@app.route("/opplan", methods=['Get', 'Post'])
def opplan():

    my_form = OpplanForm()

    # get data from database && adjust accordingly
    access_code_list = ['', 'Protected', 'Enhanced', 'Declined']
    probability_list = ['', '> 51%', '> 75%', '100%']
    tier_list = ['', 'Preferred', 'Non-Preferred', 'Specialty', 'Not Covered']
    position_list = ['', 'Exclusive', '1 of 2', '1 of 3', '1 of Many']
    restriction_list = ['', 'UR', 'PA', 'SE', 'SE/PA', '2SE']

    # header
    my_form.access_code.choices = FormUtil.get_choices(access_code_list)
    my_form.probability.choices = FormUtil.get_choices(probability_list)

    # current
    my_form.current_tier.choices = FormUtil.get_choices(tier_list)
    my_form.current_position.choices = FormUtil.get_choices(position_list)
    my_form.current_restriction.choices = FormUtil.get_choices(
        restriction_list)

    # future
    my_form.future_tier.choices = FormUtil.get_choices(tier_list)
    my_form.future_position.choices = FormUtil.get_choices(position_list)
    my_form.future_restriction.choices = FormUtil.get_choices(restriction_list)

    # print("method", request.method)

    opplan_data = OpplanData()
    opplan_data.remove_none_values()
    print("my_form.validate_on_submit()", my_form.validate_on_submit())
    print(my_form.errors)
    if my_form.validate_on_submit():
        # header
        opplan_data.change_date = request.form.get('change_date')
        opplan_data.access_code = request.form.get('access_code')
        opplan_data.probability = request.form.get('probability')

        # current
        opplan_data.current_tier = request.form.get('current_tier')
        opplan_data.current_position = request.form.get('current_position')
        opplan_data.current_restriction = request.form.get(
            'current_restriction')

        # future
        opplan_data.future_tier = request.form.get('future_tier')
        opplan_data.future_position = request.form.get('future_position')
        opplan_data.future_restriction = request.form.get('future_restriction')

        opplan_data.comments = request.form.get('comments')
        if opplan_data.comments:
            opplan_data.comments = unicodedata.normalize(
                'NFD', opplan_data.comments)

        #
        opplan_data.file_photo = ""
        file = request.files.get('file_photo')
        if file:
            filename = secure_filename(file.filename)
            # print("path", os.path.join(app.config['UPLOAD_FOLDER']))
            new_filename = str(uuid.uuid1())
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], new_filename))

            opplan_data.file_photo_filename = file.filename
            opplan_data.file_photo_code = new_filename

        opplan_data.remove_none_values()
        db.session.add(opplan_data)
        db.session.commit()
        print("****", opplan_data.id, "****")

        return redirect('/opplan/' + str(opplan_data.id))

    return render_template('opplan.html', my_form=my_form, data=opplan_data)


@app.route("/opplan/<int:id>", methods=['Get'])
def opplan_by_id(id):

    opplan_data = OpplanData.query.filter_by(id=id).first()
    print(opplan_data)
    if opplan_data == None:
        abort(404)

    return render_template('opplan2.html', data=opplan_data)


@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404


@app.errorhandler(500)
def page_not_found(e):
    return render_template('500.html'), 500


@app.route('/download/<path:filename>')
def downloadFile(filename):
    path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    return send_file(path, as_attachment=True)


@app.route("/profile", methods=['Get', 'Post'])
def profile():

    my_form = ProfileForm()

    my_data = Profile()
    my_data.remove_none_values()

    # print("my_form.validate_on_submit()", my_form.validate_on_submit())
    # print(my_form.errors)
    if my_form.validate_on_submit():

        # print("************ FORM SUBMITTED****")
        my_data.first_name = request.form.get('first_name')
        my_data.last_name = request.form.get('last_name')
        my_data.email = request.form.get('email')
        my_data.dob = request.form.get('dob')
        # print("first_name", my_data.first_name)
        # print("last_name", my_data.last_name)

        # process file
        file = request.files.get('file_photo')
        if file:
            orig_filename = secure_filename(file.filename)
            file_extension = os.path.splitext(orig_filename)
            file_extension = str(file_extension[1]).lower()

            new_filename = str(uuid.uuid1()) + file_extension

            # save to upload folder
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], new_filename))
            # print("file saved")

            my_data.file_photo_filename = orig_filename
            my_data.file_photo_code = new_filename

        # ---------------EXCEL/CSV - Load into table
        data_file = request.files.get('excel_file')
        print("data_file", data_file)
        if data_file:
            orig_filename = secure_filename(data_file.filename)
            file_extension = os.path.splitext(orig_filename)
            file_extension = str(file_extension[1]).lower()

            new_filename = str(uuid.uuid1()) + file_extension

            file_full_path = os.path.join(
                app.config['UPLOAD_FOLDER'], new_filename)

            # save to upload folder
            data_file.save(file_full_path)
            # print("file_full_path", file_full_path)
            my_data.file_data_filename = orig_filename
            my_data.file_data_code = new_filename

            # load the data in the table using pandas
            df = pd.read_csv(file_full_path)
            rest_list_raw = df.to_dict('records')
            rest_list = []
            for rest in rest_list_raw:
                my_rest = Restaurant()
                my_rest.bill = rest['bill']
                my_rest.tip = rest['tip']
                rest_list.append(my_rest)
            db.session.bulk_save_objects(rest_list)
            db.session.commit()

        # save to database
        db.session.add(my_data)
        db.session.commit()
        # print("my_data", my_data.id)

        # redirect to display page
        return redirect('/profile/' + str(my_data.id))  # profile/5

    return render_template('profile.html', my_form=my_form, my_data=my_data)


@app.route("/profile/<int:id>", methods=['Get'])
def profile_by_id(id):
    my_data = Profile.query.filter_by(id=id).first()

    restaurant_list = Restaurant.query.limit(5).all()
    restaurant_list_full = Restaurant.query.all()
    scatter_list = []
    for rest in restaurant_list_full:
        scatter_list.append(rest.to_x_y())

    # scatter_list = jsonify(scatter_list)
    # print(scatter_list)
    if my_data == None:
        abort(404)
    # print(my_data)
    return render_template('profile_view.html', my_data=my_data, restaurant_list=restaurant_list, restaurant_list_full=restaurant_list_full, scatter_list=scatter_list)


@app.route("/restaurant", methods=['Get'])
def restaurant():
    restaurant_list = []

    restaurants_count = Restaurant.query.count()
    restaurants = Restaurant.query.all()
    for restaurant in restaurants:
        restaurant_list.append(restaurant.to_dict())
    # print(restaurant_list)
    return jsonify({"draw": 1, "recordsTotal": restaurants_count, "recordsFiltered": restaurants_count, "data": restaurant_list})


@app.route("/export2")
def export():
    print("in export")
    # get all data
    restaurant_list = Restaurant.query.all()
    # print(restaurant_list)
    # create the excel workbook
    wb = Workbook()
    sheet = wb.active
    style_headline = 'Headline 1'
    style_data = 'Headline 1'

    # header
    sheet.cell(row=1, column=1).value = 'Bill'
    sheet.cell(row=1, column=1).style = style_headline
    sheet.cell(row=1, column=2).value = 'Tip'
    sheet.cell(row=1, column=2).style = style_headline
    # data
    row_index = 2
    for restaurant in restaurant_list:
        sheet.cell(row=row_index, column=1).value = restaurant.bill
        sheet.cell(row=row_index, column=1).style = style_data

        sheet.cell(row=row_index, column=2).value = restaurant.tip
        sheet.cell(row=row_index, column=2).style = style_data
        row_index += 1

    # add chart
    chart = ScatterChart()
    chart.title = "Scatter Chart"
    chart.style = 13
    chart.x_axis.title = 'Bill Amount'
    chart.y_axis.title = 'Tip Amount'

    xvalues = Reference(sheet, min_col=1, min_row=2, max_row=1001)
    for i in range(2, 3):
        values = Reference(sheet, min_col=i, min_row=1, max_row=1001)
        series = Series(values, xvalues, title_from_data=True)
        chart.series.append(series)

    sheet.add_chart(chart, "D1")

    filename = "restaurant.xlsx"
    full_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    print("full_path", full_path)
    # save the workbook
    wb.save(full_path)

    return send_file(full_path, as_attachment=True)


@app.route("/employee", methods=['Get', 'Post'])
def employee():
    my_form = EmployeeForm()
    # convert to list

    if my_form.validate_on_submit():  # my_form.submitted()
        # file we are importing
        file_csv = request.files.get('file_csv')

        if file_csv:
            file_full_path = os.path.join(
                app.config['UPLOAD_FOLDER'], file_csv.filename)
            # print("file_full_path", file_full_path)

            # save to upload folder
            file_csv.save(file_full_path)

            # load the data in the table using pandas
            df = pd.read_csv(file_full_path)

            # print("raw_data", df.iloc[0])

            # print("shape", df.shape)
            employee_list_raw = df.to_dict('records')

            # print("dictionary", employee_list_raw)

            employee_list = []
            for curr_emp in employee_list_raw:
                emp = Employee.from_dict(curr_emp)
                employee_list.append(emp)
                # db.session.add(emp)
                # db.session.commit()

            print("employee_list_count", len(employee_list))

            # save t0 DB
            db.session.bulk_save_objects(employee_list)
            db.session.commit()

            # test query
            e_list = Employee.query.limit(5).all()
            print("*******")
            print(e_list)
            print("*******")

        # send us to the display page
        # return redirect("/employee/" + str(my_data.id))

    return render_template('employee.html', my_form=my_form)


@app.route("/tables")
def tables():
    my_data = Employee.query.all()

    chart_data = []
    chart_data2 = []
    for emp in my_data:
        chart_data.append(emp.to_x_y())
        chart_data2.append(emp.to_x_y2())

    return render_template('tables_data.html', my_data=my_data, chart_data=chart_data, chart_data2=chart_data2)


@app.route("/tables.json", methods=['Get'])
def tables_api():
    e_list = []

    e_count = Employee.query.count()
    emp_all = Employee.query.all()
    for emp in emp_all:
        e_list.append(emp.to_dict())
    # print(restaurant_list)
    return jsonify({"draw": 1, "recordsTotal": e_count, "recordsFiltered": e_count, "data": e_list})


if __name__ == "__main__":
    app.run(debug=True)
